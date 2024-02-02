import psycopg2
import openpyxl
from openpyxl import load_workbook
import schedule
import time
from datetime import datetime, timedelta

# PostgreSQL 연결 정보 설정
db1_config = {
    'host': 'localhost',
    'user': 'postgres',
    'password': 'psql',
    'database': 'ncsm'
}

db2_config = {
    'host': 'localhost',
    'user': 'postgres',
    'password': 'psql',
    'database': 'ncsm'
}

db3_config = {
    'host': 'localhost',
    'user': 'postgres',
    'password': 'psql',
    'database': 'ncsm'
}

# 엑셀 파일 생성
try:
    wb = load_workbook('output.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Time', '라이브서버', 'STG서버', 'New STG(20분전)', 'New STG(30분전)', 'New STG(40분전)', 'New STG(50분전)'])  # 헤더 추가
# 엑셀 파일 저장
wb.save('output.xlsx')
#
# # PostgreSQL에 연결하고 row count 추출
# def get_row_count(config):
#     conn = psycopg2.connect(**config)
#     print("aaaa")
#     cursor = conn.cursor()
#     cursor.execute('SELECT COUNT(*) FROM common_xfactor_common')
#
#     row_count = cursor.fetchone()[0]
#     print(row_count)
#     conn.close()
#     return row_count

def get_row_count_live(config):
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()

        # 오늘 날짜와 어제 날짜 계산
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)

        cursor.execute("""
        SELECT COUNT(*), date_trunc('hour', user_date)
        FROM common_xfactor_common_cache
        WHERE user_date BETWEEN %s AND %s
        AND cache_date BETWEEN %s AND %s
        AND date_trunc('hour', user_date) = date_trunc('hour', cache_date)
        GROUP BY date_trunc('hour', user_date)
        ORDER BY date_trunc('hour', user_date);
        """, (yesterday, today, yesterday, today))
        rows = cursor.fetchall()
        row_count = rows[-1][0]

        conn.close()
        return row_count
    except Exception as e:
        print(e)


def get_row_count_stg(config):
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()

        # 오늘 날짜와 어제 날짜 계산
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)

        cursor.execute("""
        SELECT COUNT(*), date_trunc('hour', user_date)
        FROM common_xfactor_common_cache
        WHERE user_date BETWEEN %s AND %s
        AND cache_date BETWEEN %s AND %s
        AND date_trunc('hour', user_date) = date_trunc('hour', cache_date)
        GROUP BY date_trunc('hour', user_date)
        ORDER BY date_trunc('hour', user_date);
        """, (yesterday, today, yesterday, today))
        rows = cursor.fetchall()

        row_count = rows[-1][0]
        conn.close()
        return row_count
    except Exception as e:
        print(e)


def get_row_count_newstg2(config):
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()

        # 오늘 날짜와 어제 날짜 계산
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)

        cursor.execute("""
        select count(*), date_trunc('hour', user_date)
        from common_xfactor_common_cache
        WHERE user_date BETWEEN %s AND %s
        AND cache_date BETWEEN %s AND %s
        and date_trunc('hour', user_date) = date_trunc('hour', cache_date + interval '20 minutes')
        group by date_trunc('hour', user_date)
        order by date_trunc('hour', user_date);
        """, (yesterday, today, yesterday-timedelta(minutes=20), today))
        rows = cursor.fetchall()

        row_count = rows[-1][0]
        conn.close()
        return row_count
    except Exception as e:
        print(e)

def get_row_count_newstg3(config):
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()

        # 오늘 날짜와 어제 날짜 계산
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)

        cursor.execute("""
        select count(*), date_trunc('hour', user_date)
        from common_xfactor_common_cache
        WHERE user_date BETWEEN %s AND %s
        AND cache_date BETWEEN %s AND %s
        and date_trunc('hour', user_date) = date_trunc('hour', cache_date + interval '30 minutes')
        group by date_trunc('hour', user_date)
        order by date_trunc('hour', user_date);
        """, (yesterday, today, yesterday-timedelta(minutes=30), today))
        rows = cursor.fetchall()

        row_count = rows[-1][0]
        conn.close()
        return row_count
    except Exception as e:
        print(e)

def get_row_count_newstg4(config):
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()

        # 오늘 날짜와 어제 날짜 계산
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)

        cursor.execute("""
        select count(*), date_trunc('hour', user_date)
        from common_xfactor_common_cache
        WHERE user_date BETWEEN %s AND %s
        AND cache_date BETWEEN %s AND %s
        and date_trunc('hour', user_date) = date_trunc('hour', cache_date + interval '40 minutes')
        group by date_trunc('hour', user_date)
        order by date_trunc('hour', user_date);
        """, (yesterday, today, yesterday-timedelta(minutes=40), today))
        rows = cursor.fetchall()

        row_count = rows[-1][0]
        conn.close()
        return row_count
    except Exception as e:
        print(e)

def get_row_count_newstg5(config):
    try:
        conn = psycopg2.connect(**config)
        cursor = conn.cursor()

        # 오늘 날짜와 어제 날짜 계산
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday = today - timedelta(days=1)

        cursor.execute("""
        select count(*), date_trunc('hour', user_date)
        from common_xfactor_common_cache
        WHERE user_date BETWEEN %s AND %s
        AND cache_date BETWEEN %s AND %s
        and date_trunc('hour', user_date) = date_trunc('hour', cache_date + interval '50 minutes')
        group by date_trunc('hour', user_date)
        order by date_trunc('hour', user_date);
        """, (yesterday, today, yesterday-timedelta(minutes=50), today))
        rows = cursor.fetchall()

        row_count = rows[-1][0]
        conn.close()
        return row_count
    except Exception as e:
        print(e)


# 각 데이터베이스의 row count와 현재 시간을 엑셀 파일에 추가
def job():
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_count_live = get_row_count_live(db1_config)
    row_count_stg = get_row_count_stg(db2_config)
    row_count_tds2 = get_row_count_newstg2(db3_config)
    row_count_tds3 = get_row_count_newstg3(db3_config)
    row_count_tds4 = get_row_count_newstg4(db3_config)
    row_count_tds5 = get_row_count_newstg5(db3_config)

    # 기존 엑셀 파일 열기
    wb = openpyxl.load_workbook('output.xlsx')
    ws = wb.active

    # 행 추가
    ws.append([current_time, row_count_live, row_count_stg, row_count_tds2, row_count_tds3, row_count_tds4, row_count_tds5])

    # 엑셀 파일 저장
    wb.save('output.xlsx')

# # 매 시간마다 job 함수 실행
# schedule.every().hour.at(":00").do(job)
#
# # 스케줄링 시작
# while True:
#     schedule.run_pending()
#     time.sleep(1)
